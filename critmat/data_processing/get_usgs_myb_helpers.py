import os
import openpyxl
import regex as re
import numpy as np
import pandas as pd
from critmat.data_processing.translation import *
from openpyxl.styles import Alignment
'''
A collection of fuctions essential for processing the USGS Mineral Yearbook data using the get_usgs_myb function.
'''

def extract_title_info(wb,sheet):
    """Extract all title-related information from a sheet.

    Combines detect_title_cell, extract_material, extract_category, and extract_unit
    into a single function. Returns a dict with all extracted info or None if no
    valid title cell is found.
    """
    edgecases = []

    a2 = str(wb[sheet]['A2'].value)
    a3 = str(wb[sheet]['A3'].value)
    combined = a2 + a3
        
    #first a quick check if production data can be found in the sheet
    if re.search(r'WORLD.*PRODUCTION.*COUNTRY', a2, re.IGNORECASE):
        title_cell = 'A2'
        title = a2
    elif re.search(r'WORLD.*PRODUCTION.*COUNTRY', combined, re.IGNORECASE):
        title_cell = 'A2'
        title = combined
        edgecases.append('Long title')
    elif re.search(r'PRODUCTION.*WORLD', combined, re.IGNORECASE) and not re.search(r'CAPACITY', combined, re.IGNORECASE):
        title_cell = 'A2'
        title = a2
        edgecases.append('Title missing keyword')
    else:
        return None

    if found := re.search(r'(.*):', title): #The Material should be in the title like "IRON ORE:"
        title_material = found.group(1)
        if re.search(r'AND', title_material):  #Quick check if the sheet contains multiple materials
            edgecases.append('Multiple Materials')
    elif found := re.search(r'(.*), WORLD', title): #I only know of a single endgecase
        title_material = found.group(1)
        edgecases.append('Weird Title Material')
    else:
        title_material = None

    #Here we try to extract the category of production from the title, or the material name
    found = re.search(r'WORLD(.*) PRODUCTION', title)
    title_category = found.group(1) if found else None
    if not title_category:
        found = re.search(r'PRODUCTION(.*), BY COUNTRY', title)
        title_category = found.group(1) if found else None

    #This is an old implementation of the category detection currently still necessary for the database
    if re.search(r'REFINERY|PROCESSING|SECONDARY|SMELTER|ALUMINUM|ARSENIC|FERROCHROMIUM|FERROMANGANESE|PLANT|SILICON METAL', title, re.IGNORECASE):
        db_category = 'refined'
    elif re.search(r'MINE|ORE|BAUXITE|CHROMITE', title):
        db_category = 'primary'
    else:   # everything that is not explicitly refining or primary is treated as primary either way
        db_category = 'primary'

    return {
        'title_cell': title_cell,
        'title': title,
        'title_material': title_material,
        'title_category': title_category,
        'db_category': db_category,
        'edgecases': edgecases
    }


def checkExcelIntegrity(wb):
    """This function is a new implementation to precheck the loaded workbook for edgecases and details"""
    production_sheets = {}
    for sheet in wb.sheetnames:
        info = extract_title_info(wb, sheet)
        if not info: continue

        title = info['title']
        title_cell = info['title_cell']
        title_material = info['title_material']
        title_category = info['title_category']
        db_category = info['db_category']
        edgecases = info['edgecases']

        indents = []
        overflow_counter = 0 #some edgecase files have >1000 max_row, we stop the loop at 10 empty spaces in a row        
        datacolumns_cell = None
        grandtotal_cell = None
        unit_cell = None
        start_idx = int(title_cell[-1])+1

        for row in wb[sheet].iter_rows(min_row=start_idx):
            columnA_cell = row[0]
            cell_value = columnA_cell.value
            cell_value_str = str(cell_value) if cell_value is not None else ''

            #this gets the indentations used in the sheet starting after the datacolumns field. This is helpfull for the edgcase where not only country data is in the datacolumn
            if datacolumns_cell:
                try:
                    if cell_value_str.startswith("   "):
                        indents.append(1)
                    else:
                        indents.append(int(columnA_cell.alignment.indent))
                except:
                    indents.append(0)

            if cell_value:
                overflow_counter = 0 #Reset the counter if a value was found

                #This if checks for the unit usally found like "(metirc tons)"
                if (not unit_cell) and title_cell and (re.search(r'tons',str(columnA_cell.value).lower()) or re.search(r'kilogram',str(columnA_cell.value).lower()) or re.search(r'carats',str(columnA_cell.value).lower())):
                    unit_cell = columnA_cell.coordinate
                #This checks for the start of the data usally "Country..." the length of this row is used to check if all year data is present later on
                if (not datacolumns_cell) and title_cell and unit_cell and (title_cell !=  columnA_cell.coordinate) and re.search(r'countr',str(columnA_cell.value).lower()): 
                    datacolumns_cell = columnA_cell.coordinate
                    datacolumns_length = sum(1 for cell in wb[sheet][columnA_cell.row] if cell.value is not None)
                #Check to find the total end value (unused)
                if (not grandtotal_cell) and datacolumns_cell and re.search(r'grand total',str(columnA_cell.value).lower()): 
                    grandtotal_cell = columnA_cell.coordinate
                #check if we find another title cell -> this implies the existence of multiple tables
                if title in columnA_cell.value:
                    edgecases += ['Linebreaks in Table']   
                if cell_value and cell_value == 'W':
                    edgecases += ['Withheld data']   
                # #Check if last value and not including Footnotes (usally containing "Estimate" or "Revised")
                # if re.search(r'-- Zero',columnA_cell.value) or re.search(r'Estimate',columnA_cell.value) or re.search(r'Revised',columnA_cell.value) or re.search(r'rounded',columnA_cell.value): 
                #     last_row = row_idx                
                last_row = columnA_cell.row
            elif overflow_counter > 10:
                #I don't think this edgecase is noteworthy
                # edgecases += ['Max Row longer then content']
                break
            else:
                overflow_counter += 1
        #is there data at all?
        if (not title_cell) or (not datacolumns_cell): #Title check should be irrelevant now, but keeping it for safety
            continue
        #Output for edgecases:
        if datacolumns_length > 6: 
            try: 
                int_check = int(wb[sheet].cell(row=int(datacolumns_cell[-1]),column=3).value) #This is just to make sure the first column is a year
                #If there is no extra columns above the datacolumns_cell there are no extra columns (The likely cause of this is a Footnote at one of the year columns)
                if extra_columns:= [extracolumn for extracolumn in [wb[sheet].cell(row=int(datacolumns_cell[-1])-1,column=idx).value for idx in range(1,datacolumns_length*2)] if extracolumn is not None]:
                    edgecases += ['High Column number']
                else:
                    extra_columns = None
            except: 
                print('ERROR: High Column number can currently not be fixed')
                continue                           
        elif datacolumns_length < 6: 
            edgecases += ['Low Column number']
            extra_columns = None
        else:
            extra_columns = None

        #Unit edgecase
        if not unit_cell:
            print('ERROR: No unit found')
            continue
        else:
            unit_value = re.sub(r"[\n\t]*", "", str(wb[sheet][unit_cell].value)).strip() #Just removing whitespace and tabs

        #Indents edgecase. If more then half of all lines are indents something is wrong and we subtract 1 indent from everyline (see Gemstones myb)
        #We use -1 indentation later
        try:
            first_country = wb[sheet].cell(row=int(datacolumns_cell[-1])+1,column=1).value
            second_country = wb[sheet].cell(row=int(datacolumns_cell[-1])+2,column=1).value
            if first_country.lower().strip() > second_country.lower().strip():        
                indents = [x - 1 for x in indents]
                edgecases += ['Fixed indentation Edgecase']
        except:
            edgecases += ['Indentation Edgecase Check failed']
        
        #If there are extra columns the materials are likely found there
        if 'Multiple Materials' in edgecases and 'High Column number' in edgecases:
            edgecases.remove('Multiple Materials')
            edgecases.remove('High Column number')
            edgecases += ['Submaterials in extra columns']
        #If the extra materials are not in the extra columns they are likely in the indentations
        #This checks if the standart indentation handeling is used (Not the Fixed edgecase above and not the no indents case)
        if sum(indent for indent in indents) > 1:
            if 'Multiple Materials' in edgecases:
                edgecases.remove('Multiple Materials')
                edgecases += ['Submaterials in indentations']
            else:
                edgecases += ['Unkown values in indentations']
            

        if title_cell != 'A2': edgecases += ['Shifted title cell']
        if unit_cell != 'A4': edgecases += ['Shifted unit cell']
        if datacolumns_cell != 'A6': edgecases += ['Shifted data cell']

        production_sheets[sheet] = {
            'sheet': sheet, 'title_cell': title_cell, 'title_material': title_material, 
            'title_category': title_category, 'db_category': db_category, 'unit_cell': unit_cell, 
            'unit': unit_value, 'datacolumns_cell': datacolumns_cell, 'edgecases': edgecases, 
            'extra_columns': extra_columns, 'last_row': last_row, 'indents': indents
        }

    if production_sheets == []:
        print('No Production sheets found')
    return production_sheets


def readAndProcess(path,sheet,skip_rows,lastrow,edgecases):
    '''Using pandas, a sheet of interest of a file at a path is read. Columns that
    are completely empty for visual reasons automatically get the column name "Unnamed: 0,1,...
    After dropping them, some ugly conversions are needed to get all data that are not column
    or country names into a numeric data type. So random characters, empty strings, zeros all 
    get replaced with np.nan.'''
    data = pd.read_excel(path, skiprows=skip_rows, sheet_name=sheet)
    #Edgecase handeling for missing year values (see rare-earth myb 2006)
    if 'Low Column number' in edgecases:
        previous_empty = False
        fixed = False
        for i in range(1,len(data.columns)):
            if previous_empty and data.columns[i][:7] == 'Unnamed':
                data = data.rename(columns={data.columns[i]: last_entry+1})
                previous_empty = False
                fixed = True
            elif data.columns[i][:7] == 'Unnamed':
                previous_empty = True
            else:
                last_entry = int(data.columns[i][:7])
                previous_empty = False
        if not fixed:
            raise Exception('ERROR: Low Column number could not be fixed')

    dropcols = []
    for i in range(len(data.columns)):
        try: 
            if data.columns[i][:7] == 'Unnamed':
                dropcols.append(i)
        except: pass

    # CONVERT HUMAN-READABLE MESS TO NUMERIC VALUES
    data = data.drop(data.columns[dropcols],axis=1)
    pd.set_option('future.no_silent_downcasting', True)     	    # annoying. Like this I am "opting into future behavior" to suppress warnings
    data.replace(to_replace='--',value='0',inplace=True)
    data.iloc[:,1:] = data.iloc[:,1:].replace(to_replace=r'\(\d+\)',value=np.nan,regex=True)                       # remove strange symbols 
    #data.iloc[:,1:] = data.iloc[:,1:].replace(to_replace=r'W',value=999999,regex=True)                       # THIS IS TO Handel missing US DATA
    data.iloc[:,1:] = data.iloc[:,1:].replace(to_replace=r'[a-zA-Z,.\(\)\-\`]*',value='',regex=True)                       # remove strange symbols 
    data.iloc[:,1:] = data.iloc[:,1:].infer_objects(copy=False).replace(to_replace=r'^\s*$',value=np.nan,regex=True)    # NaN instead of empty strings and spaces                                                         
    data.iloc[:,1:] = data.iloc[:,1:].astype(float)     # THIS LINE CAN CAUSE ERRORS WITH PANDAS 3
    droprows = range(lastrow,data.shape[0])
    data = data.drop(index=list(droprows),axis=0)

    return data


def edgecaseCountries(liste):
    '''Quick fix that stuck around. Very often, country names come with "Continued" or "e" for
    "estimate" which felt easier to take care of explicitly rather than make edge cases in the standardize 
    function.'''
    if len(liste) == 0:
        return liste
    new_liste = []
    for i in range(len(liste)):
        word = liste[i]
        if word.endswith('Continued'):
            word = word[:-9]
        if word.endswith('Continuede'):
            word = word[:-10]
        word = word.strip()
        new_liste.append(word)
    new_liste = pd.DataFrame(new_liste)
    new_liste.columns = ['country_name']
    new_liste = standardize(new_liste,cnames=True)   # Strip the word of special characters, commas, and standardize it
    new_liste = list(new_liste['country_name'])
    liste = []
    for i in range(len(new_liste)):                         # Second loop to remove ountries that end with an "e" (for "estimate"), at least
        word = new_liste[i]                                 # if the name without the "e" is also present in the translated list. 
        if (word[-1] == 'e') and (word[:-1] in new_liste):  # So a country needs to not be an estimate just once to be recognized like this. 
            liste.append(word[:-1])
        elif (word[-1] == 'r') and (word[:-1] in new_liste):
            liste.append(word[:-1])
        else:
            liste.append(word)
    return liste

def edgecaseYears(liste):
    '''The years sometimes have an "e" in them because it is an estimation. We want numbers.'''
    new_liste = []
    for i in range(len(liste)):     
        year = str(liste[i])
        year = year.split('.',1)[0] #This removes pandas duplicate column styling (like ".1")
        year = re.sub(r'[a-zA-Z,.\(\)\-]','',year)
        year = year.strip()
        try: 
            new_liste.append(int(year))
        except:
            new_liste.append('')    # Mark strange years to be removed in format_usgsarchive or checked manually. 
    return new_liste

def handel_notes(dataframe):
    submaterial_columns = dataframe['column_note'].str.contains('content')
    if submaterial_columns.any():
        dataframe.loc[submaterial_columns,'material_name'] = dataframe.loc[submaterial_columns,'column_note'].str.split(' content',expand=True)[0].str.lstrip()
    return dataframe

def takeCareOfAllRegularMaterialsKeepInfos(df,data,sheet,indents,emptyrows,material,thisUnit,category,file_year,extra_columns):
    #This is a rework of a previously used function
    '''Rewriting logic into database format for the majority of materials. Takes the complete
    resulting df that is being constructed, the data of the current sheet and additional information
    that dedicated functions already found.'''
    # Regular materials look like this:
    #                       2000    2001    ...
    # Argentina
    # Australia
    #   submaterial 1
    #   submaterial 2
    # Brazil
    # ...
    # So we loop over all rows (skipping empty rows and "total" rows) and store an entire row as
    # five separate rows (one production value for each year) in ges (this country's total). If 
    # there is an indentation, the submaterials get summed up before being concatenated to the complete df. 
    ges = None 
    #The following lists are used to collect data rows and concat them later
    indent_rows = []
    data_rows = []

    for i in range(data.shape[0]):
        row_note = None
        column_note = None

        try: data.iloc[i,0].split(',') #this seems necessary for nan entries
        except: continue        
        row_name = data.iloc[i,0].split(',')
        if indents[i] == -1: #This is only possible in an edgecase
            material = str(data.iloc[i,0])  

        if 'Grand total' in row_name[0]:  
            break
        if re.search(r'total',str(data.iloc[i,0]),re.IGNORECASE): 
            continue # If it is "total" data, we skip the row.         

        if i in emptyrows and (i+1 >= len(indents) or indents[i+1] < 1):
            if ges:
                data_rows.append(ges)
                ges = None
            continue
        prod = [round(p) if not pd.isna(p) else 0 for p in data.iloc[i,1:].values]

        if indents[i] == 0:     # If the indentation is 0, this is the start of a new country
            country = row_name[0]
            if ges:
                data_rows.append(ges)

            if row_name[0][0].isdigit(): 
                continue            #This removes any footnotes left in the data 

            try:                 
                if 'North' in row_name[1] or 'Republic' in row_name[1] or 'Dubai' in row_name[1]: #There might be more cases
                    row_name = [str(row_name[0] + ' ' + row_name[1])]
                    row_note = ''
                else:                   
                    row_note= ' '.join(row_name[1:])    # country, year, and unit get repeated 5 times, one for each year
            except: 
                row_note = ''

            #Handel extra columns Edgecase
            column_note = pd.Series(['']*len(prod)) 
            if extra_columns:
                columnid = 0
                for column in extra_columns:
                    column_note[columnid:columnid+5] += ' ' + column
                    columnid += 5

            prod_year = list(data.iloc[i, 1:].index)
            ges = {
                'country_name': [country] * len(prod),
                'material_name': [material] * len(prod),
                'date_year': prod_year,
                'category': [category] * len(prod),
                'quantity': [0] * len(prod),
                'unit': [thisUnit] * len(prod),
                'publish_year': [file_year] * len(prod),
                'row_note': [row_note] * len(prod),
                'column_note': column_note,
                'indent_note': [''] * len(prod),
                'total_note': [''] * len(prod)
            }

            try: # if the next indent is 0 (so a new country starts), the production data must be in this row already (without submaterials)
                if (indents[i+1] == 0) or (re.search(r'total',str(data.iloc[i+1,0]),re.IGNORECASE)): 
                    ges['quantity'] = prod
            except: pass
        elif indents[i] >= 1: # if an indentation follows a non-indentation, we sum up with +=
            try:
                indent_note = str(row_note) + ' ' + str(data.iloc[i, 0]) if row_note else str(data.iloc[i, 0]) 
            except (NameError, TypeError, ValueError):
                indent_note = ''

            indent_row_df = pd.DataFrame({
                'country_name': pd.Series([country] * len(prod)), #The country name is set by indent=0
                'material_name': pd.Series([material] * len(prod)),
                'date_year': prod_year,
                'category': pd.Series([category] * len(prod)),
                'quantity': prod,
                'unit': pd.Series([thisUnit] * len(prod)),
                'publish_year': pd.Series([file_year] * len(prod)),
                'row_note': pd.Series([row_note] * len(prod)),
                'column_note': column_note,
                'indent_note': pd.Series([indent_note] * len(prod)),
                'total_note': pd.Series([''] * len(prod))
            })
            indent_rows.append(indent_row_df)
            if ges:
                ges['quantity'] = [g + p for g, p in zip(ges['quantity'], prod)]
                ges['total_note'] = [str(t) + ' ' + indent_note for t in ges['total_note']]
            else:
                continue
                # print(f'Failed to calculate Total production {country_name} of {material} row {i} in sheet {sheet}')
            # # except for one ugly edge case where Türkiye's borate refining was inflated by crude ore.
            # if material == 'boron' and country_name.iloc[0] == 'Turkey:' and not (re.search(r'refined',str(data.iloc[i,0]),re.IGNORECASE)):
            #     continue
        else: continue
    if ges:
        data_rows.append(ges)

    if data_rows:
        combined_countries = pd.concat([pd.DataFrame(c) for c in data_rows], ignore_index=True)
        df = pd.concat([df, combined_countries], ignore_index=True) if not df.empty else combined_countries

    if indent_rows:
        indent_df = pd.concat(indent_rows, ignore_index=True)
        df = pd.concat([df, indent_df], ignore_index=True)

    df = handel_notes(df)  
    return df

def prepare_for_db(dataframe):
    dataframe = dataframe[[column for column in dataframe.columns if 'note' not in column]]
    # This overwrite turns all the values ['...','...'] (old_names) into the aggregating term (key). Maybe I misunderstood but I don't think the code
    # was doing what it was supposed to before. Now, the behavior is: When we find 10 rows of Rhenium and 20 rows of Rhodium, their names get replaced by pgm
    # (so that we have 30 rows of not further specified pgms). Materials we don't want to group like this (like Palladium and Platinum) have been distinguished 
    # though edge cases earlier. 
    # material_overwrite = {} #No overwrites now
    # material_overwrite = {'Bauxite': ['Aluminium'],'Iron and Steel': ['Iron Ore'], 'Zirconium': ['Zirconium-Hafnium'],
    #                 'Rare Earths': ['Cerium','Dysprosium','Erbium','Europium','Gadolinium','Holmium','Lanthanum','Lutetium','Neodymium','Praseodymium','Samarium','Terbium','Thulium','Ytterbium','Yttrium'],
    #                 'Other platinum-group metals': ['Rhenium','Rhodium','Ruthenium','Rutile','Iridium']
    #                 }
    material_overwrite = {'aluminum primary': ['Aluminium'], 'Bauxite': ['Aluminium'], 
                          'zirconiumandhafnium': ['Zirconium','Hafnium'],
                          'industrial sand and gravel silica': ['Silica Sand'],
                          'ferrosilicon and silicon metal': ['Silicon Metal'],
                          'mined gypsum': ['Gypsum'],'LITHIUM MINERALS AND BRINE': ['Lithium'],
                          'manganese ore': ['Manganese'],'ferromanganese and silicomanganese': ['Manganse'],
                          'columbium and tantalum': ['Tantalum'],
                          'ferroniobium ferrocolumbium': ['Niobium'],
                          'niobium and tantalum': ['Niobium', 'Tantalum'],
                           'ferrochromium': ['Chromium'], 'chromite ore': ['Chromium'],
                          'quicklime and hydrated lime including deadburned dolomite': ['Lime'],
                            'rareearths': ['Cerium','Dysprosium','Erbium','Europium','Gadolinium','Holmium','Lanthanum','Lutetium','Neodymium','Praseodymium','Samarium','Terbium','Thulium','Ytterbium','Yttrium'],
                            'platinumgroup metals': ['Rhenium','Rhodium','Ruthenium','Rutile','Iridium']
                            }
    # for new_matname in material_overwrite.keys():   
    #     for old_matname in material_overwrite[new_matname]:
    #         df.loc[(df['material_name'] == old_matname),['material_name']] = new_matname # put the new name in all rows that contain the old name and in the material_name column. 
    for old_matname in material_overwrite.keys():   
        old_matdata = dataframe[dataframe['material_name']==old_matname].copy()
        for new_matname in material_overwrite[old_matname]:
            new_matdata = old_matdata.replace({'material_name': {old_matname:new_matname}}).reset_index(drop=True)
            dataframe = pd.concat([dataframe,new_matdata]).reset_index(drop=True) 
        dataframe = dataframe[dataframe['material_name']!=old_matname]
    return dataframe